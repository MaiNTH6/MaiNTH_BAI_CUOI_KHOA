# import openpyxl
# def read_excel(file_path, sheet_name="Sheet1"):
   
#     workbook = openpyxl.load_workbook(file_path)
#     sheet = workbook[sheet_name]
    
#     data = []
#     headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    
#     for row in sheet.iter_rows(min_row=2, values_only=True):
#         row_data = {headers[i]: row[i] for i in range(len(headers))}
#         data.append(row_data)
    
#     return data 
import openpyxl

def read_excel(file_path, sheet_name="Sheet1"):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    data = []
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {headers[i]: row[i] for i in range(len(headers))}

        # 🔥 FILTER: bỏ row hoàn toàn rỗng
        if not any(row_data.values()):
            continue

        data.append(row_data)

    return data
